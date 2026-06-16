import os
import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_key_pairs(data, prefix=""):
    """Recursively flattens nested dictionaries/lists to extract clean path-value pairs."""
    items = []
    if isinstance(data, dict):
        for k, v in data.items():
            full_key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, (dict, list)):
                items.extend(extract_key_pairs(v, prefix=full_key))
            else:
                items.append((full_key, str(v)))
    elif isinstance(data, list):
        for item in data:
            # Drop the list index completely to keep keys clean and structured
            if isinstance(item, (dict, list)):
                items.extend(extract_key_pairs(item, prefix=prefix))
            else:
                items.append((prefix, str(item)))
    return items

def split_path_and_key(full_path):
    """Splits a dot-notated string path into a parent path and a final target key."""
    # Clean up any residual bracket notations if they exist in the raw path strings
    clean_path = re.sub(r'\[\d+\]', '', full_path)
    
    if "." in clean_path:
        parts = clean_path.split(".")
        parent_path = ".".join(parts[:-1])
        target_key = parts[-1]
    else:
        parent_path = "Root"
        target_key = clean_path
        
    return parent_path, target_key

def build_key_pairs_excel(template_folder, output_filename="key_pairs.xlsx"):
    """Scans template JSON files and exports split path maps cleanly to Excel."""
    template_path = Path(template_folder)
    
    if not template_path.exists():
        print(f"Error: Target path '{template_path.resolve()}' does not exist.")
        return

    print(f"Reading template files from target path: {template_path.resolve()}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Template Maps"
    ws.views.sheetView[0].showGridLines = True
    
    # Updated to handle a split 4-column master layout
    headers = ["Parent Path", "Target Key", "Assigned Value", "Source File Name"]
    ws.append(headers)
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_idx == 4 else left_align
        cell.border = thin_border
    
    json_files = sorted(template_path.glob("*.json"))
    if not json_files:
        print(f"Warning: No .json files found inside '{template_path.resolve()}'")
        return
        
    for file_path in json_files:
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                raw_json = json.load(f)
            
            pairs = extract_key_pairs(raw_json)
            for key_path, val in pairs:
                parent_path, target_key = split_path_and_key(key_path)
                
                # Append row matching the new split column architecture
                ws.append([parent_path, target_key, val, file_path.name])
                
                current_row = ws.max_row
                for col_idx in range(1, 5):
                    data_cell = ws.cell(row=current_row, column=col_idx)
                    data_cell.font = Font(name="Calibri", size=10)
                    data_cell.border = thin_border
                    data_cell.alignment = center_align if col_idx == 4 else left_align
                        
        except Exception as e:
            print(f"Skipping template file {file_path.name}: {str(e)}")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)
        
    wb.save(output_filename)
    print(f"Successfully compiled all layout properties cleanly into '{output_filename}'!")

if __name__ == "__main__":
    script_directory = Path(__file__).parent
    target_templates_folder = script_directory / ".." / "templates"
    
    build_key_pairs_excel(template_folder=target_templates_folder, output_filename="key_pairs.xlsx")